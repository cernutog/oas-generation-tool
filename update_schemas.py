import openpyxl
from openpyxl.utils import get_column_letter

file_path = "API Templates/$index.xlsm"
sheet_name = "Schemas"

def update_excel():
    print(f"Opening {file_path}...")
    wb = openpyxl.load_workbook(file_path, keep_vba=True)
    ws = wb[sheet_name]

    # Find insertion point
    insert_idx = -1
    err_row_idx = -1
    
    # Iterate rows (1-based)
    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        # row is tuple. Col 0 is Name (A), Col 1 is Parent (B)
        name = row[0]
        parent = row[1]
        
        if name == "err" and parent == "ShortErrorResponse":
            err_row_idx = i
            insert_idx = i + 1 # Insert after
            print(f"Found 'err' at row {i}")
            break
    
    if insert_idx == -1:
        print("Could not find 'err' row.")
        return

    print(f"Inserting 3 rows at {insert_idx}...")
    ws.insert_rows(insert_idx, amount=3)

    # Columns (1-based indices matching the array 0-based + 1)
    # A=1 Name, B=2 Parent, C=3 Desc, D=4 Type, E=5 ItemsType, F=6 SchemaName, H=8 Mandatory, I=9 Min, J=10 Max
    
    rows_data = [
        {
            "Name": "dt", "Parent": "err", "Type": "schema", "SchemaName": "DateTime", "Mandatory": "M"
        },
        {
            "Name": "code", "Parent": "err", "Type": "string", "Desc": "A four character string representing the FPAD error code.", 
            "Min": 4, "Max": 4, "Mandatory": "M"
        },
        {
            "Name": "desc", "Parent": "err", "Type": "string", "Desc": "A string containing the error description.", "Mandatory": "M"
        }
    ]

    for i, data in enumerate(rows_data):
        r = insert_idx + i
        # Populate cells
        ws.cell(row=r, column=1, value=data["Name"])
        ws.cell(row=r, column=2, value=data["Parent"])
        if "Desc" in data: ws.cell(row=r, column=3, value=data["Desc"])
        ws.cell(row=r, column=4, value=data["Type"])
        if "SchemaName" in data: ws.cell(row=r, column=6, value=data["SchemaName"])
        if "Mandatory" in data: ws.cell(row=r, column=8, value=data["Mandatory"])
        if "Min" in data: ws.cell(row=r, column=9, value=data["Min"])
        if "Max" in data: ws.cell(row=r, column=10, value=data["Max"])

    print("Saving...")
    wb.save(file_path)
    print("Done.")

if __name__ == "__main__":
    try:
        update_excel()
    except Exception as e:
        print(f"Error: {e}")
