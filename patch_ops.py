import openpyxl
import os
import glob

base_dir = r"C:\Users\giuse\.gemini\antigravity\scratch\OAS_Generation_Tool\API Templates"

error_map = {
    "400": {"desc": "Bad Request", "schema": "ErrorResponse_400"}, 
    "401": {"desc": "Unauthorized", "schema": "ErrorResponse_401"},
    "403": {"desc": "Forbidden", "schema": "ErrorResponse_403"},
    "404": {"desc": "Not Found", "schema": "ErrorResponse_404"},
    "409": {"desc": "Conflict", "schema": "ErrorResponse_409"},
    "429": {"desc": "Too many requests", "schema": "ErrorResponse_429"},
    "500": {"desc": "Generic Error", "schema": "ErrorResponse_500"},
}

def patch_ops():
    files = glob.glob(os.path.join(base_dir, "*.xlsm"))
    for f in files:
        if "$index" in f: continue
        
        print(f"Processing {os.path.basename(f)}...")
        try:
            wb = openpyxl.load_workbook(f, keep_vba=True)
            modified = False
            
            for sheet_name in wb.sheetnames:
                if sheet_name in error_map:
                    ws = wb[sheet_name]
                    
                    headers = {}
                    header_row_idx = 1
                    # Scan first few rows to find 'Name'
                    for r_idx, row in enumerate(ws.iter_rows(max_row=5), start=1):
                        vals = [c.value for c in row]
                        if "Name" in vals:
                            header_row_idx = r_idx
                            for cell in row:
                                if cell.value: headers[cell.value] = cell.column
                            break
                    
                    if not headers:
                        print(f"  Warning: No headers found in {sheet_name}")
                        continue
                        
                    name_col = headers.get("Name")
                    type_col = headers.get("Type")
                    schema_col = headers.get("Schema Name") or headers.get("Schema Name\n(if Type = schema)") 
                    desc_col = headers.get("Description")
                    
                    info = error_map[sheet_name]
                    
                    # Target Row: First data row
                    target_row = ws[header_row_idx+1]
                    
                    # Write Data
                    # Careful not to overwrite if already correct? No, force update to fix mismatches.
                    if name_col: target_row[name_col-1].value = info["desc"]
                    if type_col: target_row[type_col-1].value = "response"
                    if schema_col and info["schema"]: target_row[schema_col-1].value = info["schema"]
                    
                    print(f"  Patched {sheet_name} with {info['desc']}")
                    modified = True
                
                elif sheet_name == "201":
                     # Patch 201 created
                     ws = wb[sheet_name]
                     headers = {}
                     header_row_idx = 1
                     for r_idx, row in enumerate(ws.iter_rows(max_row=5), start=1):
                        vals = [c.value for c in row]
                        if "Name" in vals:
                            header_row_idx = r_idx
                            for cell in row:
                                if cell.value: headers[cell.value] = cell.column
                            break
                     
                     name_col = headers.get("Name")
                     desc_col = headers.get("Description")
                     type_col = headers.get("Type")
                     schema_col = headers.get("Schema Name") or headers.get("Schema Name\n(if Type = schema)")

                     if name_col:
                         for row in ws.iter_rows(min_row=header_row_idx+1):
                             val = row[name_col-1].value
                             
                             if val == "fri":
                                 print(f"  Patching 201 fri header")
                                 if type_col: row[type_col-1].value = "header"
                                 if schema_col: row[schema_col-1].value = "FpadResponseIdentifier"
                                 modified = True
                                 
                         first_data = ws[header_row_idx+1]
                         if desc_col: first_data[desc_col-1].value = "Created"
                         modified = True

                elif sheet_name == "400":
                     ws = wb[sheet_name]
                     headers = {}
                     header_row_idx = 1
                     for r_idx, row in enumerate(ws.iter_rows(max_row=5), start=1):
                        vals = [c.value for c in row]
                        if "Name" in vals:
                            header_row_idx = r_idx
                            for cell in row:
                                if cell.value: headers[cell.value] = cell.column
                            break
                            
                     name_col = headers.get("Name")
                     type_col = headers.get("Type")
                     schema_col = headers.get("Schema Name") or headers.get("Schema Name\n(if Type = schema)")
                     desc_col = headers.get("Description")

                     if name_col:
                         for row in ws.iter_rows(min_row=header_row_idx+1):
                             val = row[name_col-1].value
                             if val == "fri":
                                 print(f"  Patching 400 fri header")
                                 if type_col: row[type_col-1].value = "header"
                                 if schema_col: row[schema_col-1].value = "FpadResponseIdentifier"
                                 modified = True
                         
                         first_data = ws[header_row_idx+1]
                         if desc_col: first_data[desc_col-1].value = "Bad Request"
                         modified = True

            if modified:
                wb.save(f)
                print(f"Saved {os.path.basename(f)}")
                
        except Exception as e:
            print(f"Error {f}: {e}")

if __name__ == "__main__":
    patch_ops()
