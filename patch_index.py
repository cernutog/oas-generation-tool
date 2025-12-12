import openpyxl
import os

base_dir = r"C:\Users\giuse\.gemini\antigravity\scratch\OAS_Generation_Tool\API Templates"
file_path = os.path.join(base_dir, "$index.xlsm")

def patch_index():
    print(f"Loading {file_path}...")
    try:
        wb = openpyxl.load_workbook(file_path, keep_vba=True)
        if "Parameters" in wb.sheetnames:
            ws = wb["Parameters"]
            headers = {cell.value: cell.column for cell in ws[1]}
            
            updates = {
                "senderBic": {"Example": "FPADITMM", "Mandatory": "Yes"},
                "pri": {"Example": "a206e009-ef37-4040-924d-db758b2950b2", "Mandatory": "Yes"}, 
                "RiskInfoArray": {"Type": "array", "Items Data Type \n(Array only)": "RiskInfo"} # Fix type
            }
            
            name_col = headers.get("Name")
            ex_col = headers.get("Example")
            mand_col = headers.get("Mandatory")
            desc_col = headers.get("Description")
            type_col = headers.get("Type")
            items_col = headers.get("Items Data Type \n(Array only)")

            if name_col:
                for row in ws.iter_rows(min_row=2):
                    name_val = row[name_col - 1].value
                    if name_val in updates:
                        print(f"Patching {name_val}...")
                        upd = updates[name_val]
                        
                        if ex_col and "Example" in upd:
                            row[ex_col - 1].value = upd["Example"]
                        if mand_col and "Mandatory" in upd:
                            row[mand_col - 1].value = upd["Mandatory"]
                        if type_col and "Type" in upd:
                            row[type_col - 1].value = upd["Type"]
                        if items_col and "Items Data Type \n(Array only)" in upd:
                            row[items_col - 1].value = upd["Items Data Type \n(Array only)"]
                            
                        if name_val == "pri" and desc_col:
                            # Remove trailing newline if present? Or just set strict text
                            curr_desc = row[desc_col - 1].value
                            if curr_desc:
                                row[desc_col - 1].value = curr_desc.strip()
            
            wb.save(file_path)
            print("Index (Parameters) patched successfully.")
            
        if "Schemas" in wb.sheetnames:
             ws = wb["Schemas"]
             headers = {cell.value: cell.column for cell in ws[1]}
             # RiskInfoArray fix
             name_col = headers.get("Name")
             type_col = headers.get("Type")
             items_col = headers.get("Items Data Type \n(Array only)")
             
             if name_col:
                 for row in ws.iter_rows(min_row=2):
                     if row[name_col-1].value == "RiskInfoArray":
                         print("Patching RiskInfoArray in Schemas...")
                         if type_col: row[type_col-1].value = "array"
                         if items_col: row[items_col-1].value = "RiskInfo"
                         break
             wb.save(file_path)
             print("Index (Schemas) patched successfully.")
             
    except Exception as e:
        print(f"Error: {e}")

if __name__ == "__main__":
    patch_index()
