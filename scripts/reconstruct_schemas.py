import yaml
import pandas as pd
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from copy import copy
import shutil
import os

GOLDEN_PATH = "Expected results/EBACL_FPAD_20251110_OpenApi3.1_FPAD_API_Participant_4.1_v20251212.yaml"
INDEX_PATH = "API Templates/$index.xlsm"
BACKUP_PATH = "API Templates/$index_backup.xlsm"

def extract_schemas_from_yaml():
    with open(GOLDEN_PATH, 'r', encoding='utf-8') as f:
        data = yaml.safe_load(f)
    return data.get("components", {}).get("schemas", {})

def flatten_schema(name, schema, parent=None, flattened_rows=[]):
    """
    Recursive function to flatten a schema definition into Excel rows.
    """
    row = {
        "Name": name,
        "Parent": parent,
        "Description": schema.get("description", ""),
        "Type": schema.get("type"),
        "Schema Name\n(if Type = schema)": None, 
        "Format": schema.get("format"),
        "Mandatory": "M" if schema.get("required_in_parent") else None,
        "Min  \nValue/Length/Item": None,
        "Max  \nValue/Length/Item": None,
        "PatternEba": None,
        "Regex": schema.get("pattern"),
        "Allowed value": ",".join([str(x) for x in schema.get("enum", [])]) if "enum" in schema else None,
        "Example": None
    }

    # Examples logic
    ex_val = []
    if "example" in schema:
        ex_val.append(str(schema["example"]))
    if "examples" in schema:
        exs = schema["examples"]
        if isinstance(exs, list):
            ex_val.extend([str(x) for x in exs])
        elif isinstance(exs, dict):
            ex_val.extend([str(v) for k, v in exs.items()])
    
    if ex_val:
        row["Example"] = " | ".join(ex_val)

    # Handle Ref
    if "$ref" in schema:
        ref_name = schema["$ref"].split("/")[-1]
        row["Schema Name\n(if Type = schema)"] = ref_name
        row["Type"] = "schema"

    # Handle Simple Constraints
    if "minLength" in schema: row["Min  \nValue/Length/Item"] = schema["minLength"]
    if "minimum" in schema: row["Min  \nValue/Length/Item"] = schema["minimum"]
    if "minItems" in schema: row["Min  \nValue/Length/Item"] = schema["minItems"]
    
    if "maxLength" in schema: row["Max  \nValue/Length/Item"] = schema["maxLength"]
    if "maximum" in schema: row["Max  \nValue/Length/Item"] = schema["maximum"]
    if "maxItems" in schema: row["Max  \nValue/Length/Item"] = schema["maxItems"]

    # Allowed Values
    if "enum" in schema:
         row["Allowed value"] = ",".join([str(x) for x in schema.get("enum", [])])

    # Handle allOf (Inheritance/Parent)
    if "allOf" in schema:
        for item in schema["allOf"]:
            if "$ref" in item:
                ref_name = item["$ref"].split("/")[-1]
                if parent is None:
                    # Root Schema inheriting from another
                    row["Parent"] = ref_name
                else:
                    # Property using allOf Ref
                    row["Type"] = "schema"
                    row["Schema Name\n(if Type = schema)"] = ref_name
            elif "type" in item or "properties" in item:
                if "properties" in item:
                     required_props = item.get("required", [])
                     for prop_name, prop_def in item["properties"].items():
                        child_def = prop_def.copy()
                        if prop_name in required_props:
                            child_def["required_in_parent"] = True
                        flatten_schema(prop_name, child_def, parent=name, flattened_rows=flattened_rows)

    # Add the row to list
    flattened_rows.append(row)

    # Handle Properties (Children)
    if "properties" in schema:
        required_props = schema.get("required", [])
        for prop_name, prop_def in schema["properties"].items():
            child_def = prop_def.copy()
            if prop_name in required_props:
                child_def["required_in_parent"] = True
            
            flatten_schema(prop_name, child_def, parent=name, flattened_rows=flattened_rows)
            
    # Handle Array items
    if str(schema.get("type")) == "array" and "items" in schema:
        items_def = schema["items"]
        if "$ref" in items_def:
             ref_name = items_def["$ref"].split("/")[-1]
             row["Items Data Type \n(Array only)"] = ref_name 
        elif items_def.get("type"):
             row["Items Data Type \n(Array only)"] = items_def["type"]

    return flattened_rows

def reconstruct():
    print("Extracting schemas from Golden Master...")
    schemas = extract_schemas_from_yaml()
    print(f"Found {len(schemas)} schemas.")
    
    all_rows = []
    for s_name, s_def in schemas.items():
        flatten_schema(s_name, s_def, parent=None, flattened_rows=all_rows)
        
    df = pd.DataFrame(all_rows)
    
    # Standard columns WITHOUT Section
    standard_cols = [
        "Name", "Parent", "Description", "Type", 
        "Items Data Type \n(Array only)", "Schema Name\n(if Type = schema)", 
        "Format", "Mandatory", "Min  \nValue/Length/Item", 
        "Max  \nValue/Length/Item", "PatternEba", "Regex", "Allowed value", "Example"
    ]
    
    for col in standard_cols:
        if col not in df.columns:
            df[col] = None
    
    df = df[standard_cols]
    
    # DO NOT BACKUP AGAIN to preserve the 'good' backup
    
    NEW_INDEX_PATH = "API Templates/$index_reconstructed.xlsm"
    print(f"Writing to {NEW_INDEX_PATH} (Sheet: Schemas)...")
    
    # Load workbook
    wb = openpyxl.load_workbook(INDEX_PATH, keep_vba=True)
    
    # Delete old Schemas sheet if exists
    if "Schemas" in wb.sheetnames:
        del wb["Schemas"]
    
    # Create new Schemas sheet
    ws = wb.create_sheet("Schemas", 0)
    
    # Write data row by row
    rows_to_write = list(dataframe_to_rows(df, index=False, header=True))
    
    for r_idx, row_data in enumerate(rows_to_write, 1):
        for c_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=r_idx, column=c_idx)
            if value is not None:
                cell.value = str(value)
            cell.number_format = '@'  # Text format
    
    # Apply header styling from screenshot
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    
    # Define column-specific header colors based on screenshot
    # Columns A-C (Name, Parent, Description): Orange/tan
    # Columns D onwards: Light gray
    orange_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
    gray_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    
    header_font = Font(name='Calibri', size=11, bold=True, color="000000")
    header_alignment = Alignment(horizontal='center', vertical='bottom', wrap_text=True)
    header_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Apply to header row
    for col_idx in range(1, len(standard_cols) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = header_border
        
        # First 3 columns get orange, rest get gray
        if col_idx <= 3:
            cell.fill = orange_fill
        else:
            cell.fill = gray_fill
    
    # Set column widths
    column_widths = {
        'A': 25,  # Name
        'B': 20,  # Parent
        'C': 35,  # Description
        'D': 12,  # Type
        'E': 18,  # Items Data Type
        'F': 20,  # Schema Name
        'G': 12,  # Format
        'H': 12,  # Mandatory
        'I': 15,  # Min
        'J': 15,  # Max
        'K': 15,  # PatternEba
        'L': 15,  # Regex
        'M': 18,  # Allowed value
        'N': 20,  # Example
    }
    
    for col_letter, width in column_widths.items():
        ws.column_dimensions[col_letter].width = width
    
    # Set tab color (teal/cyan from other component sheets)
    from openpyxl.styles.colors import Color
    ws.sheet_properties.tabColor = Color(rgb="00B0F0")
    
    # Save
    wb.save(NEW_INDEX_PATH)
    print("Done!")

if __name__ == "__main__":
    reconstruct()
